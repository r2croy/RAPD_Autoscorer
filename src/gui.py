"""
gui.py

Main GUI for RAPD Auto Scorer

Module 5

Features
--------
✓ Open RAPD Gel Image
✓ Zoom / Pan
✓ Band Selection
✓ Lane Selection
✓ Ladder Mode
✓ BP Estimation

Author : Ridoy Roy
Platform : Ubuntu
"""
import os
import csv 

from openpyxl import load_workbook

import tkinter as tk

from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import simpledialog

from .config import *

from .viewer import ImageViewer
from .image_loader import ImageLoader

from .band_picker import BandPicker
from .lane_manager import LaneManager
from .bp_estimator import BPEstimator
from .band_cluster import BandCluster
from .exporter import ReportExporter


class RAPDAutoScorerGUI:

    # ==================================================
    # Constructor
    # ==================================================

    def __init__(self):

        # ------------------------------------------
        # Main Window
        # ------------------------------------------

        self.root = tk.Tk()

        self.root.title(APP_NAME)

        self.root.geometry(
            f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}"
        )

        self.root.configure(
            bg=BACKGROUND_COLOR
        )

        # ------------------------------------------
        # Data Objects
        # ------------------------------------------

        self.loader = ImageLoader()

        self.band_picker = BandPicker()

        self.lane_manager = LaneManager()

        self.bp_estimator = BPEstimator()

        # ------------------------------------------
        # Image
        # ------------------------------------------

        self.current_image = None

        self.current_image_path = None

        # ------------------------------------------
        # Modes
        # ------------------------------------------

        self.lane_mode = False

        self.ladder_mode = False

        # First lane is ladder
        self.ladder_lane_id = 1

        # ------------------------------------------
        # Viewer
        # ------------------------------------------

        self.viewer = None

        # ------------------------------------------
        # History
        # ------------------------------------------

        self.action_history = []

        # ------------------------------------------
        # Build GUI
        # ------------------------------------------

        self.create_menu()

        self.create_toolbar()

        self.create_viewer()

        self.create_statusbar()

        self.bind_shortcuts()

        # ==========================================
        # Module-6
        # Estimated Report
        # ==========================================

        self.band_cluster = BandCluster()

        self.report_exporter = ReportExporter()

        from .binary_matrix import BinaryMatrix

        self.binary_matrix = BinaryMatrix()

    # ==================================================
    # Menu
    # ==================================================

    def create_menu(self):

        menubar = tk.Menu(
            self.root
        )

        # ------------------------------------------
        # File
        # ------------------------------------------

        file_menu = tk.Menu(
            menubar,
            tearoff=0
        )

        file_menu.add_command(

            label="Open Image",

            command=self.open_image

        )

        file_menu.add_separator()

        file_menu.add_command(

            label="Exit",

            command=self.root.destroy

        )

        menubar.add_cascade(

            label="File",

            menu=file_menu

        )

        # ------------------------------------------
        # Edit
        # ------------------------------------------

        edit_menu = tk.Menu(
            menubar,
            tearoff=0
        )

        edit_menu.add_command(

            label="Undo",

            command=self.undo_last_action

        )

        edit_menu.add_command(

            label="Reset View",

            command=self.reset_view

        )

        edit_menu.add_separator()

        edit_menu.add_command(

            label="Clear Bands",

            command=self.clear_bands

        )

        edit_menu.add_command(

            label="Clear Lanes",

            command=self.clear_lanes

        )

        menubar.add_cascade(

            label="Edit",

            menu=edit_menu

        )

        self.root.config(
            menu=menubar
        )

    # ==================================================
    # Toolbar
    # ==================================================

    def create_toolbar(self):

        self.toolbar = ttk.Frame(
            self.root
        )

        self.toolbar.pack(

            side=tk.TOP,

            fill=tk.X

        )

        # ------------------------------------------
        # Open Image
        # ------------------------------------------

        ttk.Button(

            self.toolbar,

            text="Open Image",

            command=self.open_image

        ).pack(

            side=tk.LEFT,

            padx=4,

            pady=4

        )

        # ------------------------------------------
        # Undo
        # ------------------------------------------

        ttk.Button(

            self.toolbar,

            text="Undo",

            command=self.undo_last_action

        ).pack(

            side=tk.LEFT,

            padx=4,

            pady=4

        )

        # ------------------------------------------
        # Reset
        # ------------------------------------------

        ttk.Button(

            self.toolbar,

            text="Reset",

            command=self.reset_view

        ).pack(

            side=tk.LEFT,

            padx=4,

            pady=4

        )

        # ------------------------------------------
        # Lane Mode
        # ------------------------------------------

        self.lane_button = ttk.Button(

            self.toolbar,

            text="Lane Mode : OFF",

            command=self.toggle_lane_mode

        )

        self.lane_button.pack(

            side=tk.LEFT,

            padx=8

        )

        # ------------------------------------------
        # Ladder Mode
        # ------------------------------------------

        self.ladder_button = ttk.Button(

            self.toolbar,

            text="Ladder Mode : OFF",

            command=self.toggle_ladder_mode

        )

        self.ladder_button.pack(

            side=tk.LEFT,

            padx=8

        )

        # ------------------------------------------
        # Estimate BP
        # ------------------------------------------

        ttk.Button(

            self.toolbar,

            text="Estimate BP",

            command=self.estimate_bp

        ).pack(

            side=tk.LEFT,

            padx=8

        )

        # ------------------------------------------
        # Export Binary Matrix
        # ------------------------------------------

        ttk.Button(

            self.toolbar,

            text="Export Binary",

            command=self.export_binary_matrix

        ).pack(

            side=tk.LEFT,

            padx=8

        )

        # ------------------------------------------
        # Export Estimated Report
        # ------------------------------------------

        ttk.Button(

            self.toolbar,

            text="Export Estimated Report",

            command=self.export_estimated_report

        ).pack(

            side=tk.LEFT,

            padx=8

        )

    # ==================================================
    # Viewer
    # ==================================================

    def create_viewer(self):

        frame = tk.Frame(

            self.root,

            bg="black"

        )

        frame.pack(

            fill=tk.BOTH,

            expand=True

        )

        self.viewer = ImageViewer(
            frame
        )

        # ------------------------------------------
        # Mouse Click
        # ------------------------------------------

        self.viewer.canvas.bind(

            "<Button-1>",

            self.left_click

        )

        self.viewer.canvas.bind(

            "<Button-3>",

            self.right_click

        )

    # ==================================================
    # Status Bar
    # ==================================================

    def create_statusbar(self):

        self.status_text = tk.StringVar()

        self.status_text.set(
            "Ready"
        )

        status = ttk.Label(

            self.root,

            textvariable=self.status_text,

            anchor="w"

        )

        status.pack(

            side=tk.BOTTOM,

            fill=tk.X

        )

    # ==================================================
    # Keyboard Shortcut
    # ==================================================

    def bind_shortcuts(self):

        self.root.bind(

            "<Control-o>",

            lambda event: self.open_image()

        )

        self.root.bind(

            "<Control-z>",

            lambda event: self.undo_last_action()

        )

        self.root.bind(

            "<Escape>",

            lambda event: self.reset_view()

        )

        self.root.bind(

            "<Delete>",

            lambda event: self.clear_bands()

        )

    # ==================================================
    # Status Helper
    # ==================================================

    def set_status(self, text):

        self.status_text.set(text)

    # ==================================================
    # Current Image
    # ==================================================

    def get_current_image(self):

        return self.current_image

    # ==================================================
    # Current Image Path
    # ==================================================

    def get_current_image_path(self):

        return self.current_image_path
    
    # ==================================================
    # Open Image
    # ==================================================

    def open_image(self):

        filename = filedialog.askopenfilename(

            title="Open RAPD Gel Image",

            filetypes=[

                (

                    "Image Files",

                    "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"

                )

            ]

        )

        if filename == "":
            return

        try:

            image = self.loader.load_image(
                filename
            )

        except Exception as e:

            messagebox.showerror(

                "Image Error",

                str(e)

            )

            return

        # ------------------------------------------
        # Store Image
        # ------------------------------------------

        self.current_image = image

        self.current_image_path = filename

        # ------------------------------------------
        # Reset All Data
        # ------------------------------------------

        self.band_picker.clear()

        self.lane_manager.clear()

        self.bp_estimator.clear()

        self.action_history.clear()

        # ------------------------------------------
        # Load Viewer
        # ------------------------------------------

        self.viewer.clear()

        self.viewer.load_image(image)

        self.set_status(

            "Image Loaded Successfully"

        )

    # ==================================================
    # Reset Viewer
    # ==================================================

    def reset_view(self):

        if self.current_image is None:

            return

        self.viewer.reset_view()

        self.viewer.redraw_lane_list(

            self.lane_manager.get_all()

        )

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

        self.set_status(

            "Viewer Reset"

        )

    # ==================================================
    # Refresh Viewer
    # ==================================================

    def refresh_view(self):

        if self.current_image is None:

            return

        self.viewer.refresh()

        self.viewer.redraw_lane_list(

            self.lane_manager.get_all()

        )

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

    # ==================================================
    # Clear Bands
    # ==================================================

    def clear_bands(self):

        self.band_picker.clear()

        self.viewer.clear_band_overlay()

        self.set_status(

            "All Bands Removed"

        )

    # ==================================================
    # Clear Lanes
    # ==================================================

    def clear_lanes(self):

        self.lane_manager.clear()

        self.viewer.clear_lane_overlay()

        self.set_status(

            "All Lanes Removed"

        )

    # ==================================================
    # Clear Everything
    # ==================================================

    def clear_all(self):

        self.band_picker.clear()

        self.lane_manager.clear()

        self.bp_estimator.clear()

        self.action_history.clear()

        self.viewer.clear()

        self.current_image = None

        self.current_image_path = None

        self.set_status(

            "Project Cleared"

        )

    # ==================================================
    # Toggle Lane Mode
    # ==================================================

    def toggle_lane_mode(self):

        self.lane_mode = not self.lane_mode

        # Ladder Mode OFF
        if self.lane_mode:
            self.ladder_mode = False

        if self.lane_mode:

            self.lane_button.config(
                text="Lane Mode : ON"
            )

            self.ladder_button.config(
                text="Ladder Mode : OFF"
            )

            self.set_status(
                "Lane Selection Mode"
            )

        else:

            self.lane_button.config(
                text="Lane Mode : OFF"
            )

            self.set_status(
                "Band Selection Mode"
            )

    # ==================================================
    # Toggle Ladder Mode
    # ==================================================

    def toggle_ladder_mode(self):

        self.ladder_mode = not self.ladder_mode

        # Lane Mode OFF
        if self.ladder_mode:
            self.lane_mode = False

        if self.ladder_mode:

            self.ladder_button.config(
                text="Ladder Mode : ON"
            )

            self.lane_button.config(
                text="Lane Mode : OFF"
            )

            self.set_status(
                "Click ladder bands"
            )

        else:

            self.ladder_button.config(
                text="Ladder Mode : OFF"
            )

            self.set_status(
                "Band Selection Mode"
            )

    # ==================================================
    # Left Mouse Click
    # ==================================================

    def left_click(self, event):

        if self.current_image is None:
            return

        # --------------------------------------
        # Convert Canvas -> Image Coordinate
        # --------------------------------------

        point = self.viewer.canvas_to_image(
            event.x,
            event.y
        )

        if point is None:
            return

        img_x, img_y = point

        # --------------------------------------
        # Lane Mode
        # --------------------------------------

        if self.lane_mode:

            self.add_lane(
                img_x
            )

            return

        # --------------------------------------
        # Ladder Mode
        # --------------------------------------

        if self.ladder_mode:

            self.add_ladder_bp(
                img_x,
                img_y
            )

            return

    # ==================================================
    # Right Mouse Click
    # ==================================================

    def right_click(self, event):

        if self.current_image is None:
            return

        point = self.viewer.canvas_to_image(

            event.x,

            event.y

        )

        if point is None:
            return

        img_x, img_y = point

        # --------------------------------------
        # Normal Band Picking
        # --------------------------------------

        if not self.lane_mode and not self.ladder_mode:

            self.add_band(
                img_x,
                img_y
            )

    # ==================================================
    # Add Band
    # ==================================================

    def add_band(self, img_x, img_y):

        if self.current_image is None:
            return

        # -------------------------------
        # Create Band
        # -------------------------------

        band = self.band_picker.add(
            img_x,
            img_y
        )

        # -------------------------------
        # Auto Assign Lane
        # -------------------------------

        if self.lane_manager.count() > 0:

            self.lane_manager.assign_band(
                band
            )

        else:

            band["lane"] = None

        # -------------------------------
        # History
        # -------------------------------

        self.action_history.append(

            ("band", band["id"])

        )

        # -------------------------------
        # Draw
        # -------------------------------

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

        # -------------------------------
        # Status
        # -------------------------------

        if band["lane"] is None:

            self.set_status(

                f"Band {band['id']} Added"

            )

        else:

            self.set_status(

                f"Band {band['id']} → Lane {band['lane']}"

            )

    # ==================================================
    # Add Lane
    # ==================================================

    def add_lane(self, img_x):

        lane = self.lane_manager.add(
            img_x
        )

        # -------------------------------
        # History
        # -------------------------------

        self.action_history.append(

            ("lane", lane["id"])

        )

        # -------------------------------
        # Reassign Every Band
        # -------------------------------

        self.lane_manager.assign_all(

            self.band_picker.get_all()

        )

        # -------------------------------
        # Draw Lane
        # -------------------------------

        self.viewer.redraw_lane_list(

            self.lane_manager.get_all()

        )

        # -------------------------------
        # Draw Bands Again
        # -------------------------------

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

        # -------------------------------
        # Status
        # -------------------------------

        self.set_status(

            f"Lane {lane['id']} Added"

        )

    # ==================================================
    # Undo Last Action
    # ==================================================

    def undo_last_action(self):

        if len(self.action_history) == 0:

            self.set_status(
                "Nothing to Undo"
            )

            return

        action_type, action_id = self.action_history.pop()

        # ------------------------------------------
        # Undo Band
        # ------------------------------------------

        if action_type == "band":

            self.band_picker.undo()

            self.viewer.redraw_band_list(

                self.band_picker.get_all()

            )

            self.set_status(
                "Last Band Removed"
            )

            return

        # ------------------------------------------
        # Undo Lane
        # ------------------------------------------

        if action_type == "lane":

            self.lane_manager.undo()

            # Reassign every band
            self.lane_manager.assign_all(

                self.band_picker.get_all()

            )

            self.viewer.redraw_lane_list(

                self.lane_manager.get_all()

            )

            self.viewer.redraw_band_list(

                self.band_picker.get_all()

            )

            self.set_status(
                "Last Lane Removed"
            )

            return

    # ==================================================
    # Clear Action History
    # ==================================================

    def clear_history(self):

        self.action_history.clear()

    # ==================================================
    # Total History
    # ==================================================

    def history_count(self):

        return len(
            self.action_history
        )

    # ==================================================
    # Reassign All Bands
    # ==================================================

    def reassign_all_bands(self):

        if self.lane_manager.count() == 0:
            return

        self.lane_manager.assign_all(

            self.band_picker.get_all()

        )

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

    # ==================================================
    # Refresh Everything
    # ==================================================

    def refresh_everything(self):

        self.viewer.refresh()

        self.viewer.redraw_lane_list(

            self.lane_manager.get_all()

        )

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

    # ==================================================
    # Current Project Information
    # ==================================================

    def project_summary(self):

        return {

            "bands": self.band_picker.count(),

            "lanes": self.lane_manager.count(),

            "ladder": self.bp_estimator.count()

        }
    
    # ==================================================
    # Add Ladder BP
    # ==================================================

    def add_ladder_bp(self, img_x, img_y):

        """
        Click ladder band and enter bp value.
        """

        if self.current_image is None:
            return

        # ------------------------------------------
        # Ask BP
        # ------------------------------------------

        bp = simpledialog.askfloat(

            "Ladder Band",

            "Enter Ladder Band Size (bp)",

            minvalue=1.0,

            parent=self.root

        )

        if bp is None:
            return

        # ------------------------------------------
        # Save Ladder Band
        # ------------------------------------------

        self.bp_estimator.add_ladder_band(

            img_x,

            img_y,

            bp

        )

        # ------------------------------------------
        # Draw Marker
        # ------------------------------------------

        ladder_band = {

            "id": self.bp_estimator.count(),

            "x": img_x,

            "y": img_y,

            "lane": self.ladder_lane_id,

            "bp": bp

        }

        self.viewer.draw_band(
            ladder_band
        )

        # ------------------------------------------
        # History
        # ------------------------------------------

        self.action_history.append(

            ("ladder", self.bp_estimator.count())

        )

        # ------------------------------------------
        # Status
        # ------------------------------------------

        self.set_status(

            f"Ladder Band : {bp:.0f} bp"

        )

    # ==================================================
    # Total Ladder Bands
    # ==================================================

    def total_ladder_bands(self):

        return self.bp_estimator.count()

    # ==================================================
    # Clear Ladder
    # ==================================================

    def clear_ladder(self):

        self.bp_estimator.clear()

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

        self.set_status(

            "Ladder Cleared"

        )

    # ==================================================
    # Estimate BP
    # ==================================================

    def estimate_bp(self):

        # ------------------------------------------
        # Enough ladder bands?
        # ------------------------------------------

        if self.bp_estimator.count() < 3:

            messagebox.showwarning(

                "Calibration",

                "Please select at least 3 ladder bands."

            )

            return

        # ------------------------------------------
        # Perform Calibration
        # ------------------------------------------

        success = self.bp_estimator.calibrate()

        if not success:

            messagebox.showerror(

                "Calibration",

                "Calibration failed."

            )

            return

        # ------------------------------------------
        # Estimate Every Sample Band
        # ------------------------------------------

        bands = self.bp_estimator.estimate_all(

            self.band_picker.get_all()

        )

        # ------------------------------------------
        # Refresh Viewer
        # ------------------------------------------

        self.viewer.redraw_band_list(

            bands

        )

        # ------------------------------------------
        # Calibration Info
        # ------------------------------------------

        info = self.bp_estimator.calibration_info()

        self.set_status(

            f"Calibration Complete | R² = {info['r_squared']:.4f}"

        )

        messagebox.showinfo(

            "BP Estimation",

            "BP estimation completed successfully."

        )

    # ==================================================
    # Show Calibration Summary
    # ==================================================

    def show_calibration(self):

        if not self.bp_estimator.has_calibration():

            messagebox.showwarning(

                "Calibration",

                "No calibration available."

            )

            return

        messagebox.showinfo(

            "Calibration Summary",

            self.bp_estimator.summary()

        )

    # ==================================================
    # Export Estimated Bands
    # ==================================================

    def get_estimated_bands(self):

        return self.bp_estimator.export(

            self.band_picker.get_all()

        )
    
    # ==================================================
    # Find Nearest Picked Band
    # ==================================================

    def find_nearest_band(self, x, y, tolerance=12):

        nearest = None

        nearest_distance = None

        for band in self.band_picker.get_all():

            dx = band["x"] - x

            dy = band["y"] - y

            distance = (dx * dx + dy * dy) ** 0.5

            if nearest is None:

                nearest = band

                nearest_distance = distance

            elif distance < nearest_distance:

                nearest = band

                nearest_distance = distance

        if nearest is None:

            return None

        if nearest_distance > tolerance:

            return None

        return nearest

    # ==================================================
    # Assign BP To Existing Band
    # ==================================================

    def add_ladder_bp(self, img_x, img_y):

        """
        Assign bp to an existing picked band.
        No duplicate band is created.
        """

        band = self.find_nearest_band(
            img_x,
            img_y
        )

        if band is None:

            messagebox.showwarning(

                "Band Not Found",

                "Please click an existing band."

            )

            return

        bp = simpledialog.askfloat(

            "Ladder Band",

            "Enter Band Size (bp)",

            minvalue=1.0,

            parent=self.root

        )

        if bp is None:

            return

        # -----------------------------
        # Store BP inside picked band
        # -----------------------------

        band["bp"] = bp

        # -----------------------------
        # Add to estimator
        # -----------------------------

        self.bp_estimator.add_ladder_band(

            band["x"],

            band["y"],

            bp

        )

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

        self.set_status(

            f"Ladder Band : {bp:.0f} bp"

        )

    # ==================================================
    # Undo Ladder Band
    # ==================================================

    def undo_ladder(self):

        if self.bp_estimator.count() == 0:

            self.set_status(
                "No Ladder Band To Undo"
            )

            return

        self.bp_estimator.undo()

        # Remove bp from corresponding picked band

        ladder_points = self.bp_estimator.get_all()

        ladder_coords = {

            (round(b["x"], 2), round(b["y"], 2))

            for b in ladder_points

        }

        for band in self.band_picker.get_all():

            key = (

                round(band["x"], 2),

                round(band["y"], 2)

            )

            if key not in ladder_coords:

                band.pop("bp", None)

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

        self.set_status(
            "Last Ladder Band Removed"
        )

    # ==================================================
    # Reset Calibration
    # ==================================================

    def reset_calibration(self):

        self.bp_estimator.reset_calibration()

        for band in self.band_picker.get_all():

            if band.get("lane") != self.ladder_lane_id:

                band.pop("bp", None)

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

        self.set_status(
            "Calibration Reset"
        )

    # ==================================================
    # Estimate Again
    # ==================================================

    def estimate_again(self):

        self.reset_calibration()

        self.estimate_bp()

    # ==================================================
    # Refresh Whole Project
    # ==================================================

    def refresh_project(self):

        self.viewer.refresh()

        self.viewer.redraw_lane_list(

            self.lane_manager.get_all()

        )

        self.viewer.redraw_band_list(

            self.band_picker.get_all()

        )

    # ==================================================
    # Project Statistics
    # ==================================================

    def project_statistics(self):

        return {

            "bands": self.band_picker.count(),

            "lanes": self.lane_manager.count(),

            "ladder_bands": self.bp_estimator.count(),

            "calibrated": self.bp_estimator.has_calibration()

        }
    
    # ==================================================
    # Export Estimated Report
    # ==================================================

    def export_estimated_report(self):

        if self.current_image_path is None:

            messagebox.showwarning(

                "Export",

                "Please open an image first."

            )

            return

        if not self.bp_estimator.has_calibration():

            messagebox.showwarning(

                "Export",

                "Please estimate BP first."

            )

            return

        estimated_bands = self.get_estimated_bands()
        
        print("\n========== ESTIMATED BANDS ==========")

        for band in estimated_bands:
             print(band)

        print("=====================================\n")

        if len(estimated_bands) == 0:

            messagebox.showwarning(

                "Export",

                "No estimated bands found."

            )

            return

        # ------------------------------------------
        # Build Band Cluster
        # ------------------------------------------

        self.band_cluster.load_bands(

            estimated_bands

        )

        self.band_cluster.build_grid_with_tolerance()

        table = self.band_cluster.export_table()

        # ------------------------------------------
        # Export Files
        # ------------------------------------------

        excel_file = self.report_exporter.export_estimated_excel(

            self.current_image_path,

            table

        )

        csv_file = self.report_exporter.export_estimated_csv(

            self.current_image_path,

            table

        )

        self.set_status(

            "Estimated Report Exported Successfully"

        )

        messagebox.showinfo(

            "Export Complete",

            f"Excel:\n{excel_file}\n\nCSV:\n{csv_file}"

        )

    # ==================================================
    # Export Binary Matrix
    # ==================================================

    def export_binary_matrix(self):

        """
        Select an Estimated Report
        then convert it into
        Binary Matrix.
        """

        filename = filedialog.askopenfilename(

            title="Select Estimated Report",

            initialdir="Results",

            filetypes=[

                (
                    "Estimated Report",
                    "*_Estimated_Report.xlsx *_Estimated_Report.csv"
                ),

                (
                    "Excel Files",
                    "*.xlsx"
                ),

                (
                    "CSV Files",
                    "*.csv"
                )

            ]

        )

        if filename == "":
            return

        extension = os.path.splitext(
            filename
        )[1].lower()

        if extension == ".xlsx":

            self.export_binary_from_excel(
                filename
            )

            return

        if extension == ".csv":

            self.export_binary_from_csv(
                filename
            )

            return

        messagebox.showerror(

            "Unsupported File",

            "Please select an Estimated Report."

        )   

    # ==================================================
    # Export Binary From Excel
    # ==================================================

    def export_binary_from_excel(self, filename):

        """
        Read Estimated_Report.xlsx
        and convert it into
        Binary Matrix.
        """

        try:

            workbook = load_workbook(
                filename,
                data_only=True
            )

            sheet = workbook.active

        except Exception as error:

            messagebox.showerror(
                "Excel Error",
                str(error)
            )

            return

        # ------------------------------------------
        # Read Excel Table
        # ------------------------------------------

        table = []

        for row in sheet.iter_rows(values_only=True):

            values = []

            for cell in row:

                if cell is None:

                    values.append("")

                else:

                    values.append(str(cell))

            table.append(values)

        if len(table) <= 1:

            messagebox.showwarning(

                "Estimated Report",

                "Selected file is empty."

            )

            return

        # ------------------------------------------
        # Convert Excel Table
        # ------------------------------------------

        self.build_binary_from_table(

            table,

            filename

        )

    # ==================================================
    # Export Binary From CSV
    # ==================================================

    def export_binary_from_csv(self, filename):

        """
        Read Estimated_Report.csv
        and convert it into
        Binary Matrix.
        """

        try:

            with open(

                filename,

                "r",

                newline="",

                encoding="utf-8"

            ) as file:

                reader = csv.reader(file)

                table = [

                    row

                    for row in reader

                ]

        except Exception as error:

            messagebox.showerror(

                "CSV Error",

                str(error)

            )

            return

        # ------------------------------------------
        # Empty File?
        # ------------------------------------------

        if len(table) <= 1:

            messagebox.showwarning(

                "Estimated Report",

                "Selected file is empty."

            )

            return

        # ------------------------------------------
        # Convert CSV Table
        # ------------------------------------------

        self.build_binary_from_table(

            table,

            filename

        )

    # ==================================================
    # Build Binary Matrix From Table
    # ==================================================

    def build_binary_from_table(

        self,

        table,

        filename

    ):

        self.binary_matrix.clear()

        self.binary_matrix.load_estimated_table(

            table

        )

        self.binary_matrix.build_from_estimated_report()

        group_names = [

            "Coastal",

            "Coastal",

            "Coastal",

            "Riverine",

            "Riverine",

            "Riverine",

            "Haor",

            "Haor",

            "Haor"

        ]

        sample_names = [

            "S1",

            "S2",

            "S3",

            "S4",

            "S5",

            "S6",

            "S7",

            "S8",

            "S9"

        ]

        binary_table = self.binary_matrix.export_table(

            group_names,

            sample_names

        )

        image_name = os.path.basename(

            filename

        )

        image_name = image_name.replace(

            "_Estimated_Report.xlsx",

            ""

        )

        image_name = image_name.replace(

            "_Estimated_Report.csv",

            ""

        )

        fake_image = image_name + ".jpg"

        excel = self.report_exporter.export_binary_excel(

            fake_image,

            binary_table

        )

        csv = self.report_exporter.export_binary_csv(

            fake_image,

            binary_table

        )

        messagebox.showinfo(

            "Binary Matrix",

            f"Completed.\n\nExcel:\n{excel}\n\nCSV:\n{csv}"

        )

    # ==================================================
    # Run
    # ==================================================

    def run(self):

        self.root.mainloop()
