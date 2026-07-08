"""
exporter.py

Export module for RAPD Auto Scorer.

Responsibilities
----------------
1. Export Estimated Report
2. Export Binary Matrix
3. Save XLSX
4. Save CSV
5. Auto create Results folder
6. Auto generate output filename

Author : Ridoy Roy
Platform : Ubuntu
"""

import os
import csv

from openpyxl import Workbook

from openpyxl.styles import (
    Alignment,
    Border,
    Side,
    Font
)

from openpyxl.utils import (
    get_column_letter
)


class ReportExporter:

    """
    Export estimated report and
    binary matrix into

    Excel (.xlsx)

    and

    CSV (.csv)
    """

    # --------------------------------------------------
    # Constructor
    # --------------------------------------------------

    def __init__(self):

        # Results folder

        self.output_folder = "Results"

        os.makedirs(

            self.output_folder,

            exist_ok=True

        )

        # Excel styles

        self.center = Alignment(

            horizontal="center",

            vertical="center",

            wrap_text=True

        )

        self.header_font = Font(

            bold=True

        )

        thin = Side(

            style="thin"

        )

        self.border = Border(

            left=thin,

            right=thin,

            top=thin,

            bottom=thin

        )

    # --------------------------------------------------
    # Get Image Base Name
    # --------------------------------------------------

    def image_name(self, image_path):

        """
        Example
        -------

        primer_14.jpeg

        ↓

        primer_14
        """

        filename = os.path.basename(
            image_path
        )

        name = os.path.splitext(
            filename
        )[0]

        return name

    # --------------------------------------------------
    # Estimated Report Filename
    # --------------------------------------------------
    # --------------------------------------------------
    # Binary Matrix Folder
    # --------------------------------------------------

    def binary_folder(self):

        """
        Results/Binary_Matrix
        """

        folder = os.path.join(

            self.output_folder,

            "Binary_Matrix"

        )

        os.makedirs(

            folder,

            exist_ok=True

        )

        return folder

    # --------------------------------------------------
    # Binary Matrix Excel Name
    # --------------------------------------------------

    def binary_matrix_excel_name(

        self,

        image_path

    ):

        name = self.image_name(

            image_path

        )

        return os.path.join(

            self.binary_folder(),

            f"{name}_Binary_Matrix.xlsx"

        )

    # --------------------------------------------------
    # Binary Matrix CSV Name
    # --------------------------------------------------

    def binary_matrix_csv_name(

        self,

        image_path

    ):

        name = self.image_name(

            image_path

        )

        return os.path.join(

            self.binary_folder(),

            f"{name}_Binary_Matrix.csv"

        )

    def estimated_excel_name(
        self,
        image_path
    ):

        name = self.image_name(
            image_path
        )

        return os.path.join(

            self.output_folder,

            f"{name}_Estimated_Report.xlsx"

        )

    def estimated_csv_name(
        self,
        image_path
    ):

        name = self.image_name(
            image_path
        )

        return os.path.join(

            self.output_folder,

            f"{name}_Estimated_Report.csv"

        )

  
    # --------------------------------------------------
    # Export Estimated Report (Excel)
    # --------------------------------------------------

    def export_estimated_excel(
        self,
        image_path,
        table
    ):

        """
        Export Estimated Report
        into Excel (.xlsx)

        Parameters
        ----------
        image_path : str

        table : list
            Output from
            BandCluster.export_table()

        Returns
        -------
        str
            Output Excel filename
        """

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Estimated Report"

        # ------------------------------------------
        # Write Data
        # ------------------------------------------

        for row_index, row_data in enumerate(
            table,
            start=1
        ):

            for col_index, value in enumerate(
                row_data,
                start=1
            ):

                cell = sheet.cell(
                    row=row_index,
                    column=col_index
                )

                cell.value = value

                cell.alignment = self.center

                cell.border = self.border

                if row_index == 1:

                    cell.font = self.header_font

        # ------------------------------------------
        # Column Width
        # ------------------------------------------

        total_columns = len(table[0])

        for col in range(
            1,
            total_columns + 1
        ):

            letter = get_column_letter(col)

            sheet.column_dimensions[
                letter
            ].width = 16

        # ------------------------------------------
        # Row Height
        # ------------------------------------------

        total_rows = len(table)

        sheet.row_dimensions[1].height = 24

        for row in range(
            2,
            total_rows + 1
        ):

            sheet.row_dimensions[
                row
            ].height = 34

        # ------------------------------------------
        # Beautify Sheet
        # ------------------------------------------

        self.beautify_sheet(
            sheet,
            total_rows,
            total_columns
        )

        # ------------------------------------------
        # Save Workbook
        # ------------------------------------------

        output = self.estimated_excel_name(
            image_path
        )

        workbook.save(
            output
        )

        return output
    
    # --------------------------------------------------
    # Beautify Worksheet
    # --------------------------------------------------

    def beautify_sheet(
        self,
        sheet,
        total_rows,
        total_columns
    ):

        """
        Improve worksheet appearance.
        """

        # ------------------------------------------
        # Freeze Header
        # ------------------------------------------

        sheet.freeze_panes = "A2"

        # ------------------------------------------
        # Auto Filter
        # ------------------------------------------

        sheet.auto_filter.ref = (

            f"A1:"

            f"{get_column_letter(total_columns)}"

            f"{total_rows}"

        )

        # ------------------------------------------
        # Header Height
        # ------------------------------------------

        sheet.row_dimensions[1].height = 24

        # ------------------------------------------
        # Grid Style
        # ------------------------------------------

        for row in range(

            1,

            total_rows + 1

        ):

            for col in range(

                1,

                total_columns + 1

            ):

                cell = sheet.cell(

                    row=row,

                    column=col

                )

                cell.alignment = self.center

                cell.border = self.border

    # --------------------------------------------------
    # Export Estimated Report (CSV)
    # --------------------------------------------------

    def export_estimated_csv(
        self,
        image_path,
        table
    ):

        """
        Export Estimated Report
        as CSV.
        """

        output = self.estimated_csv_name(

            image_path

        )

        with open(

            output,

            "w",

            newline="",

            encoding="utf-8"

        ) as file:

            writer = csv.writer(file)

            writer.writerows(

                table

            )

        return output
    
    # --------------------------------------------------
    # Export Binary Matrix (Excel)
    # --------------------------------------------------

    def export_binary_excel(
        self,
        image_path,
        table
    ):

        """
        Export Binary Matrix
        into Excel (.xlsx)
        """

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Binary Matrix"

        # ------------------------------------------
        # Write Table
        # ------------------------------------------

        for row_index, row_data in enumerate(
            table,
            start=1
        ):

            for col_index, value in enumerate(
                row_data,
                start=1
            ):

                cell = sheet.cell(

                    row=row_index,

                    column=col_index

                )

                cell.value = value

                cell.alignment = self.center

                cell.border = self.border

                if row_index <= 2:

                    cell.font = self.header_font

        # ------------------------------------------
        # Column Width
        # ------------------------------------------

        total_columns = len(
            table[0]
        )

        for col in range(

            1,

            total_columns + 1

        ):

            letter = get_column_letter(
                col
            )

            if col == 1:

                sheet.column_dimensions[
                    letter
                ].width = 16

            else:

                sheet.column_dimensions[
                    letter
                ].width = 10

        # ------------------------------------------
        # Row Height
        # ------------------------------------------

        total_rows = len(
            table
        )

        sheet.row_dimensions[1].height = 24
        sheet.row_dimensions[2].height = 24

        for row in range(
            3,
            total_rows + 1
        ):

            sheet.row_dimensions[
                row
            ].height = 22

        # ------------------------------------------
        # Beautify
        # ------------------------------------------

        self.beautify_sheet(

            sheet,

            total_rows,

            total_columns

        )

        # ------------------------------------------
        # Save
        # ------------------------------------------

        output = self.binary_matrix_excel_name(

            image_path

        )

        workbook.save(
            output
        )

        return output
    
    # --------------------------------------------------
    # Export Binary Matrix (CSV)
    # --------------------------------------------------

    def export_binary_csv(
        self,
        image_path,
        table
    ):

        """
        Export Binary Matrix
        into CSV (.csv)
        """

        output = self.binary_matrix_csv_name(

            image_path

        )

        with open(

            output,

            "w",

            newline="",

            encoding="utf-8"

        ) as file:

            writer = csv.writer(

                file

            )

            writer.writerows(

                table

            )

        return output
